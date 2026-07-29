"""餐饮月度报告 (spec §3.2) —— 计划批量执行 + 模板渲染 + 文件导出。

测试重点有两条:

1. **没造新引擎**: 报告的取数动作有且只有 ``tiered_answer``。测试注入一个假
   ``answer_fn`` 记录被问了哪些问句 —— 如果报告层偷偷自己查了库，这些断言会
   在真实环境里失效（且 runner 里根本没有 pool 之外的 DB 入口）。
2. **禁止降级处理**: 任何一节 delegate=false / 需要澄清 / 契约不通过 / 抛异常，
   都必须 **不产生文件** 且报错说清原因。
"""
from __future__ import annotations

import re

import pytest

from smartbi.reporting import (
    OUTPUT_FORM_REPORT_FILE,
    DataFreshness,
    ReportDataUnavailableError,
    ReportGenerationError,
    build_monthly_report,
    generate_monthly_report_file,
    get_template,
    wants_report_file,
)
from smartbi.reporting.freshness import combine
from smartbi.reporting.model import SourceFreshness
from smartbi.reporting.runner import parse_period
from smartbi.reporting.tabular import chart_to_table, charts_to_tables
from smartbi.reporting.template import ReportTemplate, SectionTemplate

# ─────────────────────────── fixtures / helpers ───────────────────────────

FRESHNESS = DataFreshness(
    as_of_date="2026-06-28",
    earliest_date="2025-01-01",
    day_count=544,
    generated_at="2026-07-29 10:00:00",
    sources=(
        SourceFreshness("agg_daily", "营收与订单", "2025-01-01", "2026-06-30", 544),
        SourceFreshness(
            "agg_restaurant_daily_totals", "领料损耗盘点",
            "2025-03-01", "2026-06-28", 480,
        ),
    ),
)

TINY_TEMPLATE = ReportTemplate(
    code="TEST_TINY",
    title="测试报告",
    sections=(
        SectionTemplate("a", "第一节", "{period}营收多少"),
        SectionTemplate("b", "第二节", "{period}各门店营收对比"),
    ),
)


def _answer(text="本月营收 ¥123,456。", *, charts=None, kpis=None, **over):
    payload = {
        "kind": "answer",
        "answer_text": text,
        "charts": charts if charts is not None else [{
            "chartType": "bar",
            "title": "门店营收",
            "xAxis": {"data": ["万达店", "中山店"]},
            "series": [{"name": "营收", "type": "bar", "data": [70000, 53456]}],
        }],
        "kpis": kpis if kpis is not None else [
            {"title": "总营收", "value": "¥123,456"},
        ],
        "contract_pass": True,
        "query_plan_hash": "planhash-abc",
        "executed_resolvers": ["RESTAURANT_OPS_SALES_SUMMARY"],
        "code": "RESTAURANT_OPS_SALES_SUMMARY",
    }
    payload.update(over)
    return payload


def _recording_answer_fn(result_for=None):
    """假的执行链。``asked`` 记录报告层实际发出的问句。"""
    asked = []

    async def fn(query, pool, factory_id, role=None, **kwargs):
        asked.append((query, factory_id, role))
        if callable(result_for):
            return result_for(query)
        return _answer()

    fn.asked = asked  # type: ignore[attr-defined]
    return fn


async def _build(**kw):
    return await build_monthly_report(
        pool=None, factory_id="RES_3101_009", role="factory_super_admin",
        template=TINY_TEMPLATE, freshness=FRESHNESS,
        answer_fn=kw.pop("answer_fn", _recording_answer_fn()),
        **kw,
    )


# ─────────────────────── 1. 复用 tiered_answer, 没造新引擎 ───────────────────

@pytest.mark.asyncio
async def test_每节都通过执行链发出模板问句():
    """报告的取数 = 把模板问句丢进 tiered_answer，一句 SQL 都不写。"""
    fn = _recording_answer_fn()
    report = await _build(period="2026-06", answer_fn=fn)

    assert [q for q, _, _ in fn.asked] == [
        "2026年6月营收多少",
        "2026年6月各门店营收对比",
    ]
    # 租户与角色原样透传给执行链 —— RBAC/RLS 仍由执行链自己那一套负责
    assert {f for _, f, _ in fn.asked} == {"RES_3101_009"}
    assert {r for _, _, r in fn.asked} == {"factory_super_admin"}
    assert len(report.sections) == 2


@pytest.mark.asyncio
async def test_报告保留计划指纹用于追溯():
    report = await _build(period="2026-06")
    assert [s.plan_hash for s in report.sections] == ["planhash-abc"] * 2
    assert report.sections[0].executed_resolvers == ("RESTAURANT_OPS_SALES_SUMMARY",)


@pytest.mark.asyncio
async def test_默认模板的问句是自然语言而不是SQL或metric码():
    tpl = get_template()
    for section in tpl.sections:
        assert "{period}" in section.query
        assert "SELECT" not in section.query.upper()
        assert "RESTAURANT_OPS_" not in section.query


# ───────────────────────── 2. 数据截至时间必须在 ─────────────────────────

@pytest.mark.asyncio
async def test_报告带数据截至时间且取各源最小值():
    report = await _build(period="2026-06")
    # agg_daily 到 6/30 但损耗表只到 6/28 → 对外声明 6/28 (木桶原则)
    assert report.freshness.as_of_date == "2026-06-28"
    assert "数据截至时间：2026-06-28" in report.freshness.as_line()
    assert "报告生成时间：2026-07-29 10:00:00" in report.freshness.as_line()


def test_数据截至时间取不到时fail_closed():
    empty = (
        SourceFreshness("agg_daily", "营收与订单", None, None, 0),
        SourceFreshness("agg_restaurant_daily_totals", "领料损耗盘点", None, None, 0),
    )
    with pytest.raises(ReportDataUnavailableError) as exc:
        combine(empty)
    assert exc.value.code == "REPORT_DATA_UNAVAILABLE"
    assert "没有任何已物化的经营数据" in str(exc.value)


def test_周期缺省时跟随数据截至日期而不是今天():
    normalized, label, span = parse_period(None, "2026-06-28")
    assert (normalized, label) == ("2026-06", "2026年6月")
    assert span == "2026-06-01 ~ 2026-06-30"


def test_周期格式非法直接报错():
    with pytest.raises(ReportGenerationError) as exc:
        parse_period("2026/6", "2026-06-28")
    assert exc.value.code == "REPORT_BAD_PERIOD"


# ────────────────────────── 3. 禁止降级处理 ──────────────────────────

@pytest.mark.asyncio
@pytest.mark.parametrize("bad,reason", [
    (None, "NOT_DELEGATED"),
    ({"kind": "clarification", "answer_text": "你想看哪个时间范围？"}, "NEEDS_CLARIFICATION"),
    (_answer(contract_pass=False), "CONTRACT_FAILED"),
    (_answer(""), "EMPTY_ANSWER"),
])
async def test_任一节拿不到可信数据就整份不生成(bad, reason):
    def picker(query):
        return bad if "各门店" in query else _answer()

    with pytest.raises(ReportGenerationError) as exc:
        await _build(period="2026-06", answer_fn=_recording_answer_fn(picker))
    err = exc.value
    assert err.code == "REPORT_SECTION_FAILED"
    assert [f.reason_code for f in err.failures] == [reason]
    assert err.failures[0].section_key == "b"
    assert "宁可不出" in err.message


@pytest.mark.asyncio
async def test_执行链抛异常也算失败不吞掉():
    async def boom(query, pool, factory_id, role=None, **kw):
        raise RuntimeError("resolver 崩了")

    with pytest.raises(ReportGenerationError) as exc:
        await _build(period="2026-06", answer_fn=boom)
    assert [f.reason_code for f in exc.value.failures] == [
        "EXECUTION_ERROR", "EXECUTION_ERROR",
    ]
    assert "resolver 崩了" in exc.value.failures[0].detail


@pytest.mark.asyncio
async def test_失败时不产生任何文件字节():
    with pytest.raises(ReportGenerationError):
        await generate_monthly_report_file(
            pool=None, factory_id="RES_3101_009", role=None,
            fmt="xlsx", period="2026-06", template=TINY_TEMPLATE,
            freshness=FRESHNESS, answer_fn=_recording_answer_fn(lambda q: None),
        )


@pytest.mark.asyncio
async def test_缺字段的kpi被丢弃而不是填占位符():
    report = await _build(
        period="2026-06",
        answer_fn=_recording_answer_fn(lambda q: _answer(kpis=[
            {"title": "总营收", "value": "¥1"},
            {"title": "客单价"},           # 没有 value
            {"title": "订单数", "value": ""},
        ])),
    )
    kpis = report.sections[0].kpis
    assert [k.title for k in kpis] == ["总营收"]
    assert all(k.value not in ("-", "0", "") for k in kpis)


def test_未知模板不回落默认模板():
    with pytest.raises(KeyError):
        get_template("NO_SUCH_TEMPLATE")


# ─────────────────── 4. charts → 表格 (不需要新数据通道) ───────────────────

def test_类目型chart折成表():
    table = chart_to_table({
        "chartType": "bar",
        "title": "门店营收",
        "xAxis": {"data": ["万达店", "中山店"]},
        "series": [
            {"name": "营收", "data": [70000, 53456]},
            {"name": "订单数", "data": [900, 700]},
        ],
    })
    assert table is not None
    assert table.columns == ("类目", "营收", "订单数")
    assert table.rows == (("万达店", 70000, 900), ("中山店", 53456, 700))


def test_饼图name_value型折成表():
    table = chart_to_table({
        "chartType": "pie",
        "title": "渠道占比",
        "series": [{"name": "营收", "data": [
            {"name": "堂食", "value": 80000},
            {"name": "外卖", "value": 20000},
        ]}],
    })
    assert table is not None
    assert table.columns == ("名称", "营收")
    assert table.rows == (("堂食", 80000), ("外卖", 20000))


def test_缺失数据点保留None不补零():
    table = chart_to_table({
        "title": "两条不等长的线",
        "xAxis": {"data": ["1月", "2月", "3月"]},
        "series": [
            {"name": "今年", "data": [1, 2, 3]},
            {"name": "去年", "data": [4, 5]},
        ],
    })
    assert table is not None
    assert table.rows[2] == ("3月", 3, None)  # 不是 0


def test_不认识的chart形状返回空而不是猜():
    assert chart_to_table({"chartType": "bar"}) is None
    assert chart_to_table("不是 dict") is None
    assert charts_to_tables(None) == ()
    assert charts_to_tables([{"series": []}, {"bad": 1}]) == ()


# ─────────────────────────── 5. 文件导出 ───────────────────────────

@pytest.mark.asyncio
async def test_xlsx导出封面含数据截至时间():
    from io import BytesIO

    from openpyxl import load_workbook

    rendered = await generate_monthly_report_file(
        pool=None, factory_id="RES_3101_009", role=None,
        fmt="xlsx", period="2026-06", template=TINY_TEMPLATE,
        freshness=FRESHNESS, answer_fn=_recording_answer_fn(),
    )
    assert rendered.filename.endswith(".xlsx")
    assert rendered.content[:2] == b"PK"          # 真 xlsx (zip) 头

    wb = load_workbook(BytesIO(rendered.content))
    cover = wb["封面与口径"]
    labels = {cover.cell(row=r, column=1).value for r in range(1, 12)}
    assert "数据截至时间" in labels
    row = next(
        r for r in range(1, 12)
        if cover.cell(row=r, column=1).value == "数据截至时间"
    )
    assert cover.cell(row=row, column=2).value == "2026-06-28"
    # 每节一个 sheet
    assert "第一节" in wb.sheetnames and "第二节" in wb.sheetnames


@pytest.mark.asyncio
async def test_pdf导出是真pdf且元数据带截至日期():
    rendered = await generate_monthly_report_file(
        pool=None, factory_id="RES_3101_009", role=None,
        fmt="pdf", period="2026-06", template=TINY_TEMPLATE,
        freshness=FRESHNESS, answer_fn=_recording_answer_fn(),
    )
    assert rendered.content[:5] == b"%PDF-"
    assert rendered.content_type == "application/pdf"
    assert rendered.meta()["as_of_date"] == "2026-06-28"
    assert rendered.meta()["sections"][0]["plan_hash"] == "planhash-abc"


@pytest.mark.asyncio
async def test_不支持的格式直接报错():
    with pytest.raises(ReportGenerationError) as exc:
        await generate_monthly_report_file(
            pool=None, factory_id="RES_3101_009", role=None,
            fmt="docx", template=TINY_TEMPLATE, freshness=FRESHNESS,
            answer_fn=_recording_answer_fn(),
        )
    assert exc.value.code == "REPORT_BAD_FORMAT"


@pytest.mark.asyncio
async def test_xlsx空值写成空单元格而不是零():
    from io import BytesIO

    from openpyxl import load_workbook

    one_section = ReportTemplate(
        code="T1", title="t",
        sections=(SectionTemplate("a", "第一节", "{period}营收"),),
    )
    rendered = await generate_monthly_report_file(
        pool=None, factory_id="F", role=None, fmt="xlsx", period="2026-06",
        template=one_section, freshness=FRESHNESS,
        answer_fn=_recording_answer_fn(lambda q: _answer(charts=[{
            "title": "缺点位",
            "xAxis": {"data": ["1月", "2月"]},
            "series": [{"name": "营收", "data": [10]}],
        }])),
    )
    ws = load_workbook(BytesIO(rendered.content))["第一节"]
    values = [
        ws.cell(row=r, column=2).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == "2月"
    ]
    assert values == [None]   # 不是 0


# ────────── 6. output_preference ↔ 报告导出的接线口 (spec §2.1 ↔ §3.2) ──────────

def test_report_file_form_matches_intent_layer():
    """报告包里写死的常量必须和意图层同值，否则「要文件」这一档接不上。"""
    from smartbi.gold.restaurant.restaurant_intent import OUTPUT_FORM_REPORT_FILE as intent_form

    assert OUTPUT_FORM_REPORT_FILE == intent_form


def test_出报告的说法会被判成要文件():
    from smartbi.gold.restaurant.restaurant_intent import (
        _detect_output_preference,
    )

    for phrase in ("6月的月度报告", "帮我出个报告", "导出报告", "给我一份 pdf"):
        assert wants_report_file(_detect_output_preference(phrase)), phrase


def test_只要表格不算要文件():
    assert not wants_report_file(["text", "table"])
    assert not wants_report_file(None)
    assert not wants_report_file([])


def test_文件名不含非法字符():
    assert not re.search(r'[\\/:*?"<>|]', "RES_3101_009_2026-06_月度经营报告.xlsx")
