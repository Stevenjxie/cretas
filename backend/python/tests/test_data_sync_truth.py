from datetime import datetime

from openpyxl import Workbook

from smartbi.api.data_sync import _extract_ar, _extract_department, _extract_sales


def _workbook(headers: list[str], values: list[object]) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "销售明细"
    sheet.append(headers)
    sheet.append(values)
    return workbook


def test_sales_only_sheet_keeps_cost_metrics_unavailable() -> None:
    workbook = _workbook(
        ["日期", "产品名", "数量", "销售金额"],
        [datetime(2026, 7, 1), "招牌菜", 2, 100],
    )

    rows, error = _extract_sales(workbook, "F006")

    assert error is None
    assert len(rows) == 1
    assert rows[0]["amount"] == 100
    assert rows[0]["cost"] is None
    assert rows[0]["profit"] is None
    assert rows[0]["gross_margin"] is None


def test_explicit_cost_can_derive_profit_and_margin() -> None:
    workbook = _workbook(
        ["日期", "产品名", "销售金额", "成本金额"],
        [datetime(2026, 7, 1), "招牌菜", 100, 100],
    )

    rows, error = _extract_sales(workbook, "F006")

    assert error is None
    assert rows[0]["cost"] == 100
    assert rows[0]["profit"] == 0
    assert rows[0]["gross_margin"] == 0


def test_explicit_percentage_is_normalized_without_default_estimate() -> None:
    workbook = _workbook(
        ["日期", "产品名", "销售金额", "毛利率"],
        [datetime(2026, 7, 1), "招牌菜", 200, 25],
    )

    rows, error = _extract_sales(workbook, "F006")

    assert error is None
    assert rows[0]["gross_margin"] == 0.25
    assert rows[0]["profit"] == 50
    assert rows[0]["cost"] == 150


def test_ar_does_not_invent_collection_amount() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "应收账款账龄"
    sheet.append(["日期", "客户", "应收余额", "回款金额"])
    sheet.append([datetime(2026, 7, 1), "客户A", 1000, None])

    rows, error = _extract_ar(workbook, "F006")

    assert error is None
    assert rows[0]["record_date"] == "2026-07-01"
    assert rows[0]["receivable_amount"] == 1000
    assert rows[0]["collection_amount"] is None


def test_department_metrics_require_explicit_period_and_do_not_invent_headcount() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "月度经营分析"
    sheet.append(["2026-07"])
    sheet.append([None, None, None, None, None, None, None, "上海分部", 100, 20, None, None])

    rows, error = _extract_department(workbook, "F006")

    assert error is None
    assert rows[0]["record_date"] == "2026-07-01"
    assert rows[0]["headcount"] is None
    assert rows[0]["sales_target"] is None
    assert rows[0]["per_capita_sales"] is None
    assert rows[0]["per_capita_cost"] is None
    assert rows[0]["cost_amount"] == 800000


def test_unstructured_income_summary_is_rejected_instead_of_estimated() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "收入及净利简表"
    sheet.append(["上海分部", 100, 20])

    rows, error = _extract_department(workbook, "F006")

    assert rows == []
    assert "未自动推造成本、人数或目标" in error
