"""
Auto-adaptation endpoint: Derive system dashboard data from uploaded Excel files.

When an Excel file is uploaded to SmartBI, this endpoint can be called to
automatically populate the system data tables (smart_bi_sales_data,
smart_bi_finance_data, smart_bi_department_data) from detected sheet types.

POST /api/smartbi/sync-system-data
{
    "upload_id": 123,
    "factory_id": "F001",
    "file_path": "/path/to/uploaded.xlsx"  // optional, if not provided reads from DB
}

The endpoint auto-detects available sheets and extracts what it can:
- 销售明细 → smart_bi_sales_data
- 利润表 sheets → smart_bi_finance_data (COST/REVENUE)
- 应收账款账龄 → smart_bi_finance_data (AR)
- 费用预算执行 → smart_bi_finance_data (BUDGET)
- 收入及净利简表 / 月度经营分析 → smart_bi_department_data
"""

import logging
import os
import re
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter()


def _get_cretas_db_url() -> str:
    """Get cretas_db connection URL (system tables are here, not smartbi_db)."""
    from config import get_settings
    settings = get_settings()
    return settings.food_kb_db_url  # This connects to cretas_db


def _safe_num(val, default=0):
    if val is None:
        return default
    try:
        f = float(val)
        return default if f != f else f
    except (ValueError, TypeError):
        return default


def _optional_num(val):
    """Parse a numeric cell without turning missing/invalid data into zero."""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    try:
        number = float(val)
        return None if number != number else number
    except (ValueError, TypeError):
        return None


def _normalize_record_date(value) -> Optional[str]:
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.isoformat()
    if value is None:
        return None
    text = str(value).strip()
    match = re.search(r'(20\d{2})[年./-](\d{1,2})(?:[月./-](\d{1,2}))?', text)
    if not match:
        return None
    year, month, day = (int(match.group(1)), int(match.group(2)), int(match.group(3) or 1))
    try:
        return date(year, month, day).isoformat()
    except ValueError:
        return None


def _find_sheet_record_date(ws) -> Optional[str]:
    """Find an explicit business period in the sheet; never use today's date."""
    for row in ws.iter_rows(
        min_row=1,
        max_row=min(ws.max_row, 10),
        min_col=1,
        max_col=min(ws.max_column, 12),
    ):
        for cell in row:
            parsed = _normalize_record_date(cell.value)
            if parsed:
                return parsed
    return None


def _safe_str(val, max_len=200):
    if val is None:
        return None
    s = str(val).replace("'", "''").strip()
    return s[:max_len] if len(s) > max_len else s


class SyncRequest(BaseModel):
    upload_id: Optional[int] = None
    factory_id: str = "F001"
    file_path: Optional[str] = None


class SyncResult(BaseModel):
    success: bool
    sales_count: int = 0
    finance_count: int = 0
    department_count: int = 0
    skipped_sheets: list = []
    errors: list = []


def _detect_sheets(wb):
    sheets = set(wb.sheetnames)
    return {
        'has_sales_detail': '销售明细' in sheets,
        'has_ar_aging': '应收账款账龄' in sheets,
        'has_monthly_kpi': '月度经营分析' in sheets,
        'has_revenue_summary': '收入及净利简表' in sheets,
        'has_budget_execution': '费用预算执行' in sheets,
        'profit_tables': [s for s in sheets if '利润表' in s],
    }


def _extract_sales(wb, factory_id):
    """Extract sales records from 销售明细 sheet."""
    if '销售明细' not in wb.sheetnames:
        return [], "销售明细 not found"

    ws = wb['销售明细']
    rows = []

    # Find header row
    header_row = None
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and '日期' in str(v) for v in vals if v):
            header_row = r
            break

    if not header_row:
        return [], "Cannot find header in 销售明细"

    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h = str(h).strip()
        if '日期' in h:
            col_map['date'] = i
        elif '客户' in h:
            col_map['customer'] = i
        elif '产品' == h or '产品名' in h:
            col_map['product'] = i
        elif '数量' in h:
            col_map['quantity'] = i
        elif '单价' in h:
            col_map['unit_price'] = i
        elif '毛利率' in h:
            col_map['gross_margin'] = i
        elif '成本' in h:
            col_map['cost'] = i
        elif '利润' in h or '毛利额' in h:
            col_map['profit'] = i
        elif '金额' in h:
            col_map['amount'] = i
        elif '区域' in h:
            col_map['region'] = i
        elif '业务员' in h or '销售' in h:
            col_map['salesperson'] = i

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    for r in range(header_row + 1, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(ws.max_column)]
        if not any(v is not None for v in vals):
            continue
        first = vals[0]
        if first and isinstance(first, str) and ('合计' in first or '总计' in first):
            continue

        order_date = vals[col_map.get('date', 0)] if 'date' in col_map else None
        if order_date is None:
            continue

        if isinstance(order_date, datetime):
            date_str = order_date.strftime('%Y-%m-%d')
        elif hasattr(order_date, 'isoformat'):
            date_str = order_date.isoformat()
        else:
            date_str = str(order_date).strip()

        customer = _safe_str(vals[col_map['customer']] if 'customer' in col_map else None)
        product = _safe_str(vals[col_map['product']] if 'product' in col_map else None, 100)
        quantity = _safe_num(vals[col_map['quantity']] if 'quantity' in col_map else 0)
        unit_price = _safe_num(vals[col_map['unit_price']] if 'unit_price' in col_map else 0)
        amount = _safe_num(vals[col_map['amount']] if 'amount' in col_map else 0)
        region = _safe_str(vals[col_map['region']] if 'region' in col_map else None, 100)
        salesperson = _safe_str(vals[col_map['salesperson']] if 'salesperson' in col_map else None, 100)

        cost = _optional_num(vals[col_map['cost']]) if 'cost' in col_map else None
        profit = _optional_num(vals[col_map['profit']]) if 'profit' in col_map else None
        gross_margin = (
            _optional_num(vals[col_map['gross_margin']])
            if 'gross_margin' in col_map
            else None
        )
        if gross_margin is not None and abs(gross_margin) > 1:
            gross_margin /= 100

        # Only derive one financial field from another value supplied by the
        # workbook. A sales-only sheet has no defensible cost basis, so keep all
        # three fields NULL instead of silently assuming a 67% cost ratio.
        if cost is not None and profit is None:
            profit = amount - cost
        elif profit is not None and cost is None:
            cost = amount - profit
        elif gross_margin is not None and cost is None and profit is None:
            profit = amount * gross_margin
            cost = amount - profit

        if gross_margin is None and profit is not None and amount != 0:
            gross_margin = profit / amount

        cost = round(cost, 2) if cost is not None else None
        profit = round(profit, 2) if profit is not None else None
        gross_margin = round(gross_margin, 4) if gross_margin is not None else None

        rows.append({
            'factory_id': factory_id,
            'order_date': date_str,
            'salesperson_name': salesperson,
            'region': region,
            'customer_name': customer,
            'product_name': product,
            'product_category': product,
            'quantity': quantity,
            'amount': amount,
            'unit_price': unit_price,
            'cost': cost,
            'profit': profit,
            'gross_margin': gross_margin,
            'created_at': now,
            'updated_at': now,
        })

    return rows, None


def _extract_ar(wb, factory_id):
    """Extract AR aging records."""
    if '应收账款账龄' not in wb.sheetnames:
        return [], "应收账款账龄 not found"

    ws = wb['应收账款账龄']
    rows = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Find header with '客户' and detect aging bucket columns
    header_row = None
    for r in range(1, min(5, ws.max_row + 1)):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v and '客户' in str(v) for v in vals if v):
            header_row = r
            break

    if not header_row:
        return [], "Cannot find header in 应收账款账龄"

    # Multi-row header scan
    col_map = {}
    for scan_row in [header_row, header_row + 1]:
        for i in range(ws.max_column):
            h = ws.cell(scan_row, i + 1).value
            if h is None:
                continue
            h = str(h).strip()
            if '客户' in h:
                col_map['customer'] = i
            elif '日期' in h or '期间' in h:
                col_map['date'] = i
            elif '回款' in h or '已收' in h:
                col_map['collection'] = i
            elif '应收' in h:
                col_map['balance'] = i
            elif '0-30' in h:
                col_map['d0_30'] = i
            elif '31-60' in h:
                col_map['d31_60'] = i
            elif '61-90' in h:
                col_map['d61_90'] = i
            elif '91-180' in h:
                col_map['d91_180'] = i
            elif '180天以上' in h:
                col_map['d180plus'] = i

    data_start = header_row + 2 if 'd0_30' in col_map else header_row + 1

    for r in range(data_start, ws.max_row + 1):
        vals = [ws.cell(r, c + 1).value for c in range(ws.max_column)]
        customer = vals[col_map.get('customer', 0)] if 'customer' in col_map else None
        if customer is None or (isinstance(customer, str) and ('合计' in customer or '占比' in customer)):
            continue

        record_date = (
            _normalize_record_date(vals[col_map['date']])
            if 'date' in col_map
            else _find_sheet_record_date(ws)
        )
        if record_date is None:
            continue

        balance = _safe_num(vals[col_map.get('balance', 1)] if 'balance' in col_map else 0)
        if balance <= 0:
            continue

        d0_30 = _safe_num(vals[col_map.get('d0_30', 2)] if 'd0_30' in col_map else 0)
        d31_60 = _safe_num(vals[col_map.get('d31_60', 3)] if 'd31_60' in col_map else 0)
        d61_90 = _safe_num(vals[col_map.get('d61_90', 4)] if 'd61_90' in col_map else 0)
        d91_180 = _safe_num(vals[col_map.get('d91_180', 5)] if 'd91_180' in col_map else 0)
        d180plus = _safe_num(vals[col_map.get('d180plus', 6)] if 'd180plus' in col_map else 0)

        total_weighted = d0_30 * 15 + d31_60 * 45 + d61_90 * 75 + d91_180 * 135 + d180plus * 270
        avg_aging = int(total_weighted / balance) if balance > 0 else 0
        collection = (
            _optional_num(vals[col_map['collection']])
            if 'collection' in col_map
            else None
        )

        rows.append({
            'factory_id': factory_id,
            'record_date': record_date,
            'record_type': 'AR',
            'customer_name': _safe_str(customer),
            'receivable_amount': balance,
            'collection_amount': collection,
            'aging_days': avg_aging,
            'created_at': now,
            'updated_at': now,
        })

    return rows, None


def _extract_department(wb, factory_id):
    """Extract department data from 月度经营分析 or 收入及净利简表."""
    rows = []
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if '月度经营分析' in wb.sheetnames:
        ws = wb['月度经营分析']
        record_date = _find_sheet_record_date(ws)
        if record_date is None:
            return [], "月度经营分析缺少明确期间，未同步部门指标"
        for r in range(1, ws.max_row + 1):
            for check_col in [8, 1]:
                val = ws.cell(r, check_col).value
                if val is None:
                    continue
                val_str = str(val).strip()
                if '分部' in val_str or '区域' in val_str or '省区' in val_str:
                    if check_col == 8:
                        revenue_raw = _optional_num(ws.cell(r, 9).value)
                        net_profit_raw = _optional_num(ws.cell(r, 10).value)
                        achievement = _optional_num(ws.cell(r, 12).value)
                    else:
                        revenue_raw = _optional_num(ws.cell(r, 2).value)
                        net_profit_raw = _optional_num(ws.cell(r, 3).value)
                        achievement = _optional_num(ws.cell(r, 5).value)

                    if revenue_raw is None or revenue_raw <= 0:
                        continue
                    revenue = revenue_raw * 10000
                    net_profit = net_profit_raw * 10000 if net_profit_raw is not None else None

                    dept_name = val_str
                    if any(d['department'] == dept_name for d in rows):
                        continue

                    cost = round(revenue - net_profit, 2) if net_profit is not None else None
                    sales_target = (
                        round(revenue / achievement, 2)
                        if achievement is not None and achievement > 0
                        else None
                    )

                    rows.append({
                        'factory_id': factory_id,
                        'record_date': record_date,
                        'department': dept_name,
                        'headcount': None,
                        'sales_amount': revenue,
                        'sales_target': sales_target,
                        'cost_amount': cost,
                        'per_capita_sales': None,
                        'per_capita_cost': None,
                        'created_at': now,
                        'updated_at': now,
                    })

    if rows:
        return rows, None

    # Fallback to 收入及净利简表
    if '收入及净利简表' in wb.sheetnames:
        return [], "收入及净利简表缺少受控列映射，未自动推造成本、人数或目标"

    return rows, "No suitable sheets" if not rows else None


def _persist_to_db(db_url: str, factory_id: str, sales, ar_rows, dept_rows):
    """Write extracted data to cretas_db system tables."""
    from sqlalchemy import create_engine, text
    from sqlalchemy.orm import sessionmaker
    engine = create_engine(db_url, pool_pre_ping=True)
    Session = sessionmaker(bind=engine)
    session = Session()

    errors = []
    sales_count = 0
    finance_count = 0
    dept_count = 0

    try:
        # Replace only domains for which this workbook produced verified rows.
        # AR synchronization must never erase governed REVENUE/COST records.
        if sales:
            session.execute(
                text("DELETE FROM smart_bi_sales_data WHERE factory_id = :fid"),
                {'fid': factory_id},
            )
        if ar_rows:
            session.execute(
                text(
                    "DELETE FROM smart_bi_finance_data "
                    "WHERE factory_id = :fid AND record_type = 'AR'"
                ),
                {'fid': factory_id},
            )
        if dept_rows:
            session.execute(
                text("DELETE FROM smart_bi_department_data WHERE factory_id = :fid"),
                {'fid': factory_id},
            )

        # Insert sales
        for row in sales:
            try:
                session.execute(text("""
                    INSERT INTO smart_bi_sales_data
                    (factory_id, order_date, salesperson_name, region, customer_name,
                     product_name, product_category, quantity, amount, unit_price,
                     cost, profit, gross_margin, created_at, updated_at)
                    VALUES (:factory_id, :order_date, :salesperson_name, :region, :customer_name,
                            :product_name, :product_category, :quantity, :amount, :unit_price,
                            :cost, :profit, :gross_margin, :created_at, :updated_at)
                """), row)
                sales_count += 1
            except Exception as e:
                logger.warning(f"Sales insert error: {e}")
                errors.append("Sales insert error")

        # Insert AR
        for row in ar_rows:
            try:
                session.execute(text("""
                    INSERT INTO smart_bi_finance_data
                    (factory_id, record_date, record_type, customer_name,
                     receivable_amount, collection_amount, aging_days, created_at, updated_at)
                    VALUES (:factory_id, :record_date, :record_type, :customer_name,
                            :receivable_amount, :collection_amount, :aging_days, :created_at, :updated_at)
                """), row)
                finance_count += 1
            except Exception as e:
                logger.warning(f"AR insert error: {e}")
                errors.append("AR insert error")

        # Insert departments
        for row in dept_rows:
            try:
                session.execute(text("""
                    INSERT INTO smart_bi_department_data
                    (factory_id, record_date, department, headcount, sales_amount, sales_target,
                     cost_amount, per_capita_sales, per_capita_cost, created_at, updated_at)
                    VALUES (:factory_id, :record_date, :department, :headcount, :sales_amount, :sales_target,
                            :cost_amount, :per_capita_sales, :per_capita_cost, :created_at, :updated_at)
                """), row)
                dept_count += 1
            except Exception as e:
                logger.warning(f"Dept insert error: {e}")
                errors.append("Dept insert error")

        session.commit()
        logger.info(f"Sync complete: sales={sales_count}, finance={finance_count}, dept={dept_count}")

    except Exception as e:
        session.rollback()
        logger.error(f"Sync transaction failed: {e}")
        errors.append("Transaction error")
    finally:
        session.close()
        engine.dispose()

    return sales_count, finance_count, dept_count, errors


@router.post("/sync-system-data", response_model=SyncResult)
async def sync_system_data(req: SyncRequest):
    """
    Auto-adaptation endpoint: derive system dashboard data from uploaded Excel.

    Detects available sheets and extracts:
    - 销售明细 → smart_bi_sales_data
    - 应收账款账龄 → smart_bi_finance_data (AR)
    - 月度经营分析/收入及净利简表 → smart_bi_department_data

    Sheets that don't exist are skipped gracefully.
    """
    logger.info(f"Sync request: factory={req.factory_id}, upload={req.upload_id}, file={req.file_path}")

    if not req.file_path:
        return SyncResult(success=False, errors=["file_path is required"])

    # A1: Path traversal prevention — resolve symlinks and restrict to allowed directories
    ALLOWED_DIRS = ["/www/wwwroot/cretas/", "/tmp/smartbi-uploads/"]
    real_path = os.path.realpath(req.file_path)
    if not any(real_path.startswith(d) for d in ALLOWED_DIRS):
        logger.warning(f"Path traversal attempt blocked: {req.file_path} -> {real_path}")
        return SyncResult(success=False, errors=["File path not allowed"])

    if not os.path.exists(real_path):
        return SyncResult(success=False, errors=["File not found"])

    try:
        import openpyxl
        wb = openpyxl.load_workbook(real_path, data_only=True)
    except Exception as e:
        logger.error(f"Cannot open xlsx: {e}")
        return SyncResult(success=False, errors=["Cannot open xlsx file"])

    caps = _detect_sheets(wb)  # noqa: F841
    skipped = []

    # Extract sales
    sales, err = _extract_sales(wb, req.factory_id)
    if err:
        skipped.append(f"Sales: {err}")

    # Extract AR
    ar_rows, err = _extract_ar(wb, req.factory_id)
    if err:
        skipped.append(f"AR: {err}")

    # Extract departments
    dept_rows, err = _extract_department(wb, req.factory_id)
    if err:
        skipped.append(f"Department: {err}")

    # Skip if nothing to sync
    if not sales and not ar_rows and not dept_rows:
        return SyncResult(
            success=True,
            skipped_sheets=skipped,
            errors=["No extractable data found in this file"]
        )

    # Persist to cretas_db
    try:
        db_url = _get_cretas_db_url()
    except Exception as e:
        logger.error(f"Cannot get DB URL: {e}")
        return SyncResult(success=False, errors=["Database connection error"])

    sales_count, finance_count, dept_count, errors = _persist_to_db(
        db_url, req.factory_id, sales, ar_rows, dept_rows
    )

    return SyncResult(
        success=len(errors) == 0,
        sales_count=sales_count,
        finance_count=finance_count,
        department_count=dept_count,
        skipped_sheets=skipped,
        errors=errors,
    )
