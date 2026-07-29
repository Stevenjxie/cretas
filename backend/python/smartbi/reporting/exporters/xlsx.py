"""xlsx 渲染 (openpyxl)。

版式:
* 第一个 sheet ``封面与口径``: 标题 / 周期 / **数据截至时间** / 逐源明细 /
  每节对应的原始问句与 plan_hash（可追溯性）。
* 之后每节一个 sheet: 结论文字 → KPI → 表格。

⛔ 空值写成**空单元格**，不写 0、不写 "-"。Excel 里 0 会被 SUM 进去，"-" 会
被当成文本，两者都会把「没有这个数」悄悄变成一个数。
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any, List

from ..model import MonthlyReport, ReportSection

_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _sheet_title(raw: str, used: List[str]) -> str:
    name = _INVALID_SHEET_CHARS.sub("", raw).strip() or "Sheet"
    name = name[:28]
    candidate = name
    i = 2
    while candidate in used:
        candidate = f"{name[:26]}_{i}"
        i += 1
    used.append(candidate)
    return candidate


def _write_section(ws, section: ReportSection, bold, wrap) -> None:
    row = 1
    ws.cell(row=row, column=1, value=section.heading).font = bold
    row += 1
    ws.cell(row=row, column=1, value=f"原始问句：{section.query}")
    row += 1
    if section.plan_hash:
        ws.cell(row=row, column=1, value=f"计划指纹 plan_hash：{section.plan_hash}")
        row += 1
    row += 1

    ws.cell(row=row, column=1, value="结论").font = bold
    row += 1
    cell = ws.cell(row=row, column=1, value=section.answer_text)
    cell.alignment = wrap
    row += 2

    if section.kpis:
        ws.cell(row=row, column=1, value="关键指标").font = bold
        row += 1
        for col, kpi in enumerate(section.kpis, start=1):
            ws.cell(row=row, column=col, value=kpi.title).font = bold
            ws.cell(row=row + 1, column=col, value=kpi.value)
        row += 3

    for table in section.tables:
        ws.cell(row=row, column=1, value=table.title).font = bold
        row += 1
        for col, header in enumerate(table.columns, start=1):
            ws.cell(row=row, column=col, value=header).font = bold
        row += 1
        for data_row in table.rows:
            for col, value in enumerate(data_row, start=1):
                # None -> 空单元格 (openpyxl 默认行为), 刻意不替换成 0/"-"
                ws.cell(row=row, column=col, value=value)
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 42
    for letter in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[letter].width = 18


def render_xlsx(report: MonthlyReport) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    used: List[str] = []
    cover = wb.active
    cover.title = _sheet_title("封面与口径", used)

    rows: List[Any] = [
        (report.title, None),
        ("报告周期", report.period_label),
        ("周期区间", report.meta.get("period_span")),
        ("工厂/租户", report.factory_id),
        ("模板", report.template_code),
        # ⬇⬇ 数据截至时间: xlsx 里就在这一行 (封面与口径 sheet, A7/B7)
        ("数据截至时间", report.freshness.as_of_date),
        ("最早数据日期", report.freshness.earliest_date),
        ("覆盖天数", report.freshness.day_count),
        ("报告生成时间", report.freshness.generated_at),
    ]
    r = 1
    for label, value in rows:
        cover.cell(row=r, column=1, value=label).font = bold
        if value is not None:
            cover.cell(row=r, column=2, value=value)
        r += 1
    r += 1
    cover.cell(row=r, column=1, value="逐数据源截至情况").font = bold
    r += 1
    for header_col, header in enumerate(
        ("数据源", "说明", "最早日期", "数据截至日期", "覆盖天数"), start=1,
    ):
        cover.cell(row=r, column=header_col, value=header).font = bold
    r += 1
    for src in report.freshness.sources:
        cover.cell(row=r, column=1, value=src.source)
        cover.cell(row=r, column=2, value=src.label)
        cover.cell(row=r, column=3, value=src.min_date)
        cover.cell(row=r, column=4, value=src.max_date)
        cover.cell(row=r, column=5, value=src.day_count)
        r += 1
    r += 1
    cover.cell(row=r, column=1, value="本报告目录（每节 = 一次真实执行的查询计划）").font = bold
    r += 1
    for header_col, header in enumerate(
        ("节", "标题", "原始问句", "计划指纹", "执行 resolver"), start=1,
    ):
        cover.cell(row=r, column=header_col, value=header).font = bold
    r += 1
    for section in report.sections:
        cover.cell(row=r, column=1, value=section.key)
        cover.cell(row=r, column=2, value=section.heading)
        cover.cell(row=r, column=3, value=section.query)
        cover.cell(row=r, column=4, value=section.plan_hash)
        cover.cell(row=r, column=5, value=", ".join(section.executed_resolvers))
        r += 1

    cover.column_dimensions["A"].width = 22
    cover.column_dimensions["B"].width = 26
    cover.column_dimensions["C"].width = 40
    cover.column_dimensions["D"].width = 26
    cover.column_dimensions["E"].width = 28

    for section in report.sections:
        ws = wb.create_sheet(title=_sheet_title(section.heading, used))
        _write_section(ws, section, bold, wrap)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
